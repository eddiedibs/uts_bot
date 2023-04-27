#!/home/edd1e/scripts/projs/uts_bot/uts_bot_env/bin/python3
import configuration as conf

from time import sleep
import os
import logging
import calendar
import locale


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import WebDriverException
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s [%(levelname)s|%(name)s|%(funcName)s]:: %(message)s', 
    handlers=[logging.StreamHandler()]
)
class Helpers:


    def __init__(self, driver):
        self.driver = driver



    def create_calendar(self, month:int, year:int, wb:object):
        # Establecer la configuración regional en español
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

        # Create a new sheet for the current month
        ws = wb.create_sheet(calendar.month_name[month])

        # Set the title of the sheet to the month
        ws.title = calendar.month_name[month]

        # Establecer el idioma local a inglés
        locale.setlocale(locale.LC_TIME, 'en_US.UTF-8')

        # Set the column widths
        for col in range(1, 8):
            ws.column_dimensions[calendar.day_name[col-1][:3]].width = 15

        # Set the first row to be the day names
        for col in range(1, 8):
            ws.cell(row=1, column=col, value=calendar.day_name[col-1][:3].upper())
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

        # Get the number of days in May and the day of the week that May 1 falls on
        num_days = calendar.monthrange(year, month)[1]
        day_one = calendar.weekday(year, month, 1)

        # Start writing the dates to the worksheet
        row = 2
        col = day_one + 1
        for day in range(1, num_days+1):
            ws.cell(row=row, column=col, value=day)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            col += 1
            if col == 8:
                col = 1
                row += 1

        # Save the workbook
        wb.save(f'{year}_Calendar.xlsx')
        logging.info(f"Calendar for {calendar.month_name[month]} {year} created successfully!")
        return wb

    def check_calendar(self, month:int,year:int):
        # Load the existing workbook if it exists, otherwise create a new one
        try:
            wb = load_workbook(f"{year}_Calendar.xlsx")

            # check if sheet name already exists
            sheet_name = calendar.month_name[month]
            for sheet in wb:
                if sheet.title == sheet_name:
                    raise ValueError(f"Duplicate name found: {sheet.title}")

            return self.create_calendar(month, year, wb)

        except FileNotFoundError:
            # Create a new workbook
            logging.warning("NO BOOK FOUND... CREATING NEW...")
            wb = Workbook()
            return self.create_calendar(month, year, wb)
            # ws = wb.active


    def write_activity(self, wb:object, month_name:str, subject_data:str, year:int, day:int):
        sheet = wb[month_name]
        # iterate over all cells in the worksheet
        logging.info(f"iterating over all cells in the worksheet")
        for row in sheet.iter_rows():
            for cell in row:
                # check if the cell contains the string str(day)
                if cell.value and str(day) in str(cell.value).lower():
                    # print(f"Found cell with '{str(day)}' at", cell.coordinate)
                    cell.value = str(cell.value) + f'{subject_data}'
                    lines = str(cell.value).count('\n') + 1
                    font_size = cell.font.size
                    height = (font_size * 1.2) * lines  # assuming line height is 20% larger than font size
                    sheet.row_dimensions[cell.row].height = height
                    sheet.column_dimensions[cell.column_letter].width = len(str(cell.value))
                    break


            else:
                continue
            break
        
        # Save the changes to the file
        wb.save(f'{year}_Calendar.xlsx')


    def write_data_to_excel(self, date_string, subject_data):
        # Establecer la configuración regional en español
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

        # Get date and time from input string
        date_time = datetime.strptime(date_string, "%A, %d de %B de %Y, %H:%M")
        month = date_time.month
        day = date_time.day
        year = date_time.year

        month_name = calendar.month_name[month]
        # Load the Excel file
        try:
            wb = load_workbook(f'{year}_Calendar.xlsx')

        except FileNotFoundError:
            # Create a new workbook
            logging.warning("NO BOOK FOUND... CREATING NEW...")
            wb = Workbook()
            self.create_calendar(month, year, wb)

        # Check if the "month" sheet exists
        if month_name in wb.sheetnames:
            self.write_activity(wb, month_name, subject_data, year, day)

        else:
            new_wb = self.check_calendar(month, year)
            self.write_activity(new_wb, month_name, subject_data, year, day)
            # # Save the changes to the file
            # wb.save(f'{year}_Calendar.xlsx')


    def find_activity(self, i:int, course_sections:list):
        activities_len = len(course_sections[i].find_elements(By.CLASS_NAME, "aalink.stretched-link"))
        for index in range(activities_len):
            course_sections = self.driver.find_inner_element(By.CLASS_NAME,"course-content-item-content", is_single_element=False)
            activity_title = course_sections[i].find_elements(By.CLASS_NAME,"text-uppercase.small")
            activities = course_sections[i].find_elements(By.CLASS_NAME, "aalink.stretched-link")
            
            if activity_title[index].text.strip() in conf.UNDESIRED_ACTIVITIES:
                pass

            else:
                activity_found = self.driver.click_button(activities[index])
                if activity_found:

                    description = self.driver.find_inner_element(By.CLASS_NAME,"description-inner", wait_for_element=False)
                    if description != None:
                        item_to_remove = self.driver.find_inner_element(By.CLASS_NAME,"description-inner").find_elements(By.TAG_NAME, "div")[-1].find_element(By.TAG_NAME, "strong").text.strip()
                        activity_header = self.driver.find_inner_element(By.CLASS_NAME,"page-header-headings").text.strip().replace("\n", " ")
                        subject_title = self.driver.find_inner_element(By.CLASS_NAME,"breadcrumb-item").find_element(By.TAG_NAME, "a").text.strip()
                        if item_to_remove in description.text.strip():
                            description_text = description.find_elements(By.TAG_NAME, "div")[-1].text.strip()
                            new_description_text = description_text[len(f"{item_to_remove} "):]
                            logging.info(f"\n\nSubject:: {subject_title}\nActivity name:: {activity_header}\nScheduled for:: {new_description_text}\n\n")
                            logging.info(f"Writing to file...")
                            self.write_data_to_excel(new_description_text, f"\n\n{subject_title}::{activity_header}\n")
                            sleep(1)
                            self.driver.go_back()
                    else:
                        logging.warning(f"DEADLINE NOT FOUND ON CURRENT ACTIVITY, REDIRECTING...")
                        self.driver.go_back()

                    
                else:
                    logging.error(f"ELEMENT:: '{activities[index].text}' was not found")
                    exit()

    def get_saia_activities(self, element:object):
        # Navigate back and handle errors
        self.driver.click_button(element.find_element(By.TAG_NAME, "a"))
        course_sections_len = len(self.driver.find_inner_element(By.CLASS_NAME,"course-content-item-content", is_single_element=False))
        for i in range(1, course_sections_len):
            course_sections = self.driver.find_inner_element(By.CLASS_NAME,"course-content-item-content", is_single_element=False)
            course_classes = course_sections[i].get_attribute("class")
            # check if the "collapse show" class is present in the list of classes
            if "collapse show" in course_classes:
                # the "collapse show" class is present
                self.find_activity(i, course_sections)

            else:
                # the "collapse show" class is not present
                course_headers = self.driver.find_inner_element(By.CLASS_NAME,"course-section-header", is_single_element=False)
                self.driver.click_button(course_headers[i].find_element(By.TAG_NAME, "a"))
                # the "collapse show" class is present
                self.find_activity(i, course_sections)
                





if __name__ == "__main__":
    pass